from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import Settings
from app.models import Company
from app.serializers import as_aware


HEADERS = [
    "Company name",
    "INN",
    "OGRN",
    "Primary OKVED",
    "Additional OKVED codes",
    "Email",
    "Email source",
    "Discovery date",
    "Discovery time",
    "Status",
]


def build_xlsx(companies: list[Company], settings: Settings) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FuelLead"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{max(len(companies) + 1, 2)}"

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="EA580C")
        cell.alignment = Alignment(vertical="center")

    for company in companies:
        discovered = as_aware(company.first_discovered_at).astimezone(settings.timezone)
        primary = " — ".join(
            part for part in [company.primary_okved_code, company.primary_okved_name] if part
        )
        additional = "\n".join(
            " — ".join(part for part in [item.code, item.name] if part)
            for item in sorted(company.additional_okveds, key=lambda item: item.code)
        )
        emails = "\n".join(item.email for item in company.emails)
        sources = "\n".join(sorted({item.source for item in company.emails}))
        sheet.append(
            [
                company.name,
                company.inn,
                company.ogrn or "",
                primary,
                additional,
                emails,
                sources,
                discovered.date().isoformat(),
                discovered.strftime("%H:%M"),
                company.status,
            ]
        )

    widths = [34, 14, 17, 44, 56, 34, 24, 16, 12, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()

